import sqlite3
import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import subprocess
import json

# ===============================================================
# TIME CONSTANTS
# ===============================================================
am_start = datetime.time(7, 30, 0)  # 7:30 am start time
am_late = datetime.time(9, 30, 0)  # 9:30 am late start
am_absent = datetime.time(10, 0, 0)  # 10:15 am absent
am_end = datetime.time(14, 0, 0)  # 2:00 pm out time
am_latest_out = datetime.time(15, 0, 0)  # 3:00 pm latest out

pm_start = datetime.time(15, 1, 0)  # 3:01 pm start time
pm_late = datetime.time(16, 0, 0)  # 4:00 pm late start
pm_absent = datetime.time(16, 31, 0)  # 4:30 pm absent
pm_end = datetime.time(21, 0, 0)  # 9:00 pm out time
pm_latest_out = datetime.time(23, 59, 0)  # 11:59 pm latest out

config_file = "config.json"


# ===============================================================
# CONFIGURATION MANAGEMENT
# ===============================================================

# Load configuration from file or return defaults
def load_config():
    try:
        with open(config_file, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {
            "db_path": "C:/Program Files (x86)/ZKBio Time.Net/TimeNet.db",
            "report_directory": "C:/Users/Public/Documents",
            "department_salaries": {
                "Dining 1": 12750.0,
                "Dining 2": 12300.0,
                "Chief Cook": 28000.0,
                "Senior Cook": 25000.0,
                "Cook": 20000.0,
                "Chief Cutter": 27000.0,
                "Senior Cutter": 18000.0,
                "Cutter": 13000.0,
                "Quality Control": 16000.0,
                "Helper 1": 13000.0,
                "Helper 2": 12300.0,
                "Washer": 12300.0
            }
        }


# Get salary configuration for a specific department
def get_salary_config(dept_name):
    """Return salary configuration based on department role"""
    config = load_config()

    # Check if department exists in config
    if dept_name not in config["department_salaries"]:
        return None

    # Get department salary
    base_salary = config["department_salaries"].get(dept_name, 0)

    # Calculate all values from base salary using the same formula for all departments
    return {
        "daily_salary": base_salary / 30,
        "deduction_per_minute": ((base_salary / 30) / 8) / 60,
        "absence_deduction": (base_salary / 30) / 2
    }


# ===============================================================
# MAIN PROCESSING FUNCTION
# ===============================================================

# Process attendance data for a date range and generate Excel report
def process_dates(start_date, end_date, excel_filename):
    try:
        # Load configuration
        config = load_config()
        db_path = config["db_path"]
        report_directory = config["report_directory"]

        start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d")

        num_days = (end_dt - start_dt).days + 1
        total_shifts = num_days * 2

        # Connect to database and retrieve punch data
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            query = """
                SELECT em.id, em.emp_firstname, em.emp_lastname, em.department_id, dep.dept_name, ap.punch_time
                FROM hr_employee em
                INNER JOIN hr_department dep ON em.department_id = dep.id
                LEFT JOIN att_punches ap ON em.id = ap.employee_id
                WHERE (date(ap.punch_time) BETWEEN ? AND ?) AND (em.emp_privilege=0) AND (em.emp_active=1)
                ORDER BY em.id, ap.punch_time;
            """
            cursor.execute(query, (start_date, end_date))
            results = cursor.fetchall()

            # Initialize employee attendance dictionary
            employee_attendance = {}
            for emp_id, first_name, last_name, department_id, dept_name, punch_time_str in results:
                punch_time = datetime.datetime.strptime(punch_time_str, "%Y-%m-%d %H:%M:%S")

                # Create employee record if doesn't exist
                if emp_id not in employee_attendance:
                    # Get salary configuration based on department
                    salary_config = get_salary_config(dept_name)
                    daily_salary = salary_config["daily_salary"]
                    gross_salary = daily_salary * num_days

                    employee_attendance[emp_id] = {
                        "first_name": first_name,
                        "last_name": last_name,
                        "department_id": department_id,
                        "dept_name": dept_name,
                        "daily_salary": daily_salary,
                        "total_shifts": total_shifts,
                        "late": 0,
                        "absent": 0,
                        "gross_salary": gross_salary,
                        "punches": []
                    }

                # Add punch time to employee's record
                employee_attendance[emp_id]["punches"].append(punch_time)

            # Process attendance for each employee
            for emp_id, data in employee_attendance.items():
                # Check attendance status
                status = check_attendance(
                    data["punches"],
                    start_date,
                    end_date
                )

                # Update employee record with attendance status
                data["late"] = status["Late Minutes"]
                data["absent"] = status["Absent"]

            # Create report directory if it doesn't exist
            if not os.path.exists(report_directory):
                os.makedirs(report_directory)

            # Generate Excel report
            generate_excel(excel_filename, employee_attendance, start_date, end_date)
            print(f"Excel report generated successfully: {os.path.join(report_directory, excel_filename)}")

    except sqlite3.Error as e:
        print(f"Database error: {e}")
    except ValueError as e:
        print(f"Invalid input: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


# ===============================================================
# ATTENDANCE CHECKING
# ===============================================================

# Check attendance for an employee within the date range
def check_attendance(punches, start_date, end_date):
    status = {"Late Minutes": 0, "Absent": 0}
    punches.sort()

    # Generate a list of all dates in the range
    start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
    end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
    all_dates = [start_dt + datetime.timedelta(days=i) for i in range((end_dt - start_dt).days + 1)]

    # Process each date in the range
    for current_date in all_dates:
        am_shift = []
        pm_shift = []

        # Filter punches for the current date and split by shift
        for punch in punches:
            if punch.date() == current_date:
                if am_start <= punch.time() <= am_latest_out:
                    am_shift.append(punch)
                elif pm_start <= punch.time() <= pm_latest_out:
                    pm_shift.append(punch)

        # Check morning shift attendance
        punch_in_am = next((p for p in am_shift if am_start <= p.time() < am_absent), None)
        punch_out_am = next((p for p in am_shift if am_end <= p.time() <= am_latest_out), None)

        if punch_in_am and punch_out_am:
            # Employee was present but check if late
            if am_late <= punch_in_am.time() < am_absent:
                late_minutes = (datetime.datetime.combine(current_date, punch_in_am.time()) - datetime.datetime.combine(
                    current_date, am_late)).total_seconds() // 60
                status["Late Minutes"] += int(late_minutes)
        else:
            # Employee was absent for morning shift
            status["Absent"] += 1

        # Check afternoon shift attendance
        punch_in_pm = next((p for p in pm_shift if pm_start <= p.time() < pm_absent), None)
        punch_out_pm = next((p for p in pm_shift if pm_end <= p.time() <= pm_latest_out), None)

        if punch_in_pm and punch_out_pm:
            # Employee was present but check if late
            if pm_late <= punch_in_pm.time() < pm_absent:
                late_minutes = (datetime.datetime.combine(current_date, punch_in_pm.time()) - datetime.datetime.combine(
                    current_date, pm_late)).total_seconds() // 60
                status["Late Minutes"] += int(late_minutes)
        else:
            # Employee was absent for afternoon shift
            status["Absent"] += 1

    return status


# ===============================================================
# EXCEL REPORT GENERATION
# ===============================================================

def get_daily_attendance_status(punches, current_date):
    """
    Determine attendance status for morning and afternoon shifts for a specific date
    Returns: [morning_status, afternoon_status]
    where status is:
    '✓' for present and on time
    '#' for present but late
    '✕' for absent
    """
    am_shift = []
    pm_shift = []

    # Filter punches for the current date and split by shift
    for punch in punches:
        if punch.date() == current_date:
            if am_start <= punch.time() <= am_latest_out:
                am_shift.append(punch)
            elif pm_start <= punch.time() <= pm_latest_out:
                pm_shift.append(punch)

    # Check morning shift attendance
    punch_in_am = next((p for p in am_shift if am_start <= p.time() < am_absent), None)
    punch_out_am = next((p for p in am_shift if am_end <= p.time() <= am_latest_out), None)

    if punch_in_am and punch_out_am:
        if am_late <= punch_in_am.time() < am_absent:
            am_status = '#'  # Late
        else:
            am_status = '✓'  # Present and on time
    else:
        am_status = '✕'  # Absent

    # Check afternoon shift attendance
    punch_in_pm = next((p for p in pm_shift if pm_start <= p.time() < pm_absent), None)
    punch_out_pm = next((p for p in pm_shift if pm_end <= p.time() <= pm_latest_out), None)

    if punch_in_pm and punch_out_pm:
        if pm_late <= punch_in_pm.time() < pm_absent:
            pm_status = '#'  # Late
        else:
            pm_status = '✓'  # Present and on time
    else:
        pm_status = '✕'  # Absent

    return [am_status, pm_status]


# Generate the Excel report
def generate_excel(filename, employee_attendance, start_date, end_date):
    # Load config before generating excel
    config = load_config()
    report_directory = config["report_directory"]

    workbook = openpyxl.Workbook()

    # Rename the default sheet to 'Attendance'
    salary_sheet = workbook.active
    salary_sheet.title = "Attendance"

    # Freeze panes
    salary_sheet.freeze_panes = 'B2'  # Freeze columns

    # Create a new sheet for attendance
    attendance_sheet = workbook.create_sheet(title="Daily Attendance")

    # ------------------------------------------------------
    # FIRST SHEET: SALARY REPORT
    # ------------------------------------------------------

    # Create a horizontal-only border style (top and bottom only)
    horizontal_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths for better readability
    column_widths = {
        1: 5,  # Number
        2: 25,  # Full Name
        3: 15,  # Position
        4: 12,  # Daily
        5: 15,  # Monthly
        6: 13,  # Late Mins.
        7: 13,  # Absences
        8: 15,  # Attendance
    }

    # Apply column widths
    for col_num, width in column_widths.items():
        salary_sheet.column_dimensions[get_column_letter(col_num)].width = width

    # Create headers with formatting
    headers = ["", "NAME", "POSITION", "DAILY", "MONTHLY", "LATE MINS.", "ABSENCES", "ATTENDANCE"]

    header_row = 1
    for col_num, header_text in enumerate(headers, 1):
        cell = salary_sheet.cell(row=header_row, column=col_num, value=header_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical="center")
        cell.border = horizontal_border

    # Add employee data starting from row 2
    row_num = 2
    for i, (emp_id, data) in enumerate(employee_attendance.items(), 1):
        # Map column data
        daily_salary = data["daily_salary"]
        late_minutes = data["late"]
        absences = data["absent"] / 2
        monthly_salary = daily_salary * 30.00
        full_name = data["last_name"] + ", " + data["first_name"]

        # Add the basic data cells
        salary_sheet.cell(row=row_num, column=1, value=f"{i}.")
        salary_sheet.cell(row=row_num, column=2, value=full_name)
        salary_sheet.cell(row=row_num, column=3, value=data["dept_name"])

        # Daily Salary - editable with initial calculated value
        daily_salary_cell = salary_sheet.cell(row=row_num, column=4, value=daily_salary)
        daily_salary_cell.number_format = '#,##0.00'
        daily_salary_cell.border = horizontal_border

        # Semi-Monthly - calculated based on Daily Salary
        monthly_salary_cell = salary_sheet.cell(row=row_num, column=5, value=monthly_salary)
        monthly_salary_cell.number_format = '#,##0.00'
        monthly_salary_cell.border = horizontal_border

        # Late Minutes - initially calculated
        late_minutes_cell = salary_sheet.cell(row=row_num, column=6, value=late_minutes)
        late_minutes_cell.border = horizontal_border

        # Days Absent (initial value plus editable formula)
        absence_cell = salary_sheet.cell(row=row_num, column=7, value=absences)
        absence_cell.number_format = '#,##0.0'
        absence_cell.border = horizontal_border

        # Attendance (formula that references the absence cell)
        attendance_formula = f"=15-G{row_num}"  # Assuming column 7 is G
        attendance_cell = salary_sheet.cell(row=row_num, column=8)
        attendance_cell.value = attendance_formula  # Set the formula as the value
        attendance_cell.number_format = '#,##0.0'
        attendance_cell.border = horizontal_border
        attendance_cell.fill = PatternFill(start_color="81A8FC", end_color="81A8FC", fill_type="solid")

        row_num += 1

    for row in range(2, row_num):
        salary_sheet.row_dimensions[row].height = 22.5

    salary_sheet.row_dimensions[header_row].height = 26

    for row in range(1, row_num):
        for col in range(1, 8):
            salary_sheet.cell(row=row, column=col).border = horizontal_border

    # -------------------------------------------------------
    # SECOND SHEET: DAILY ATTENDANCE TABLE
    # -------------------------------------------------------

    # Generate date list for the selected range
    start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
    end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
    date_list = [start_dt + datetime.timedelta(days=i) for i in range((end_dt - start_dt).days + 1)]

    # Create title for attendance sheet
    attendance_title = f"Daily Attendance Record ({start_date} - {end_date})"
    title_cell = attendance_sheet.cell(row=1, column=1, value=attendance_title)
    title_cell.font = Font(size=20, bold=True)
    # Merge cells based on number of dates (3 base columns + number of date columns)
    end_column = 3 + len(date_list)
    attendance_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    title_cell.alignment = Alignment(horizontal='center')

    # Legend for symbols
    legend_row = 2
    legend_text = "✓ = Present and on time      # = Present but late      ✕ = Absent for shift"
    legend_cell = attendance_sheet.cell(row=legend_row, column=1, value=legend_text)
    legend_cell.font = Font(bold=True)
    attendance_sheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=end_column)
    legend_cell.alignment = Alignment(horizontal='center')

    # Create attendance table headers
    attendance_header_row = 4

    # Base headers: ID, First Name, Last Name
    base_headers = ["ID", "First Name", "Last Name"]

    # Create one column for each date in the range with format MM/DD
    date_headers = [date.strftime("%m/%d") for date in date_list]

    # Combine all headers
    attendance_headers = base_headers + date_headers

    # Set uniform column widths for all columns in attendance sheet
    # Base columns (ID, First Name, Last Name)
    attendance_sheet.column_dimensions[get_column_letter(1)].width = 8  # ID
    attendance_sheet.column_dimensions[get_column_letter(2)].width = 18  # First Name
    attendance_sheet.column_dimensions[get_column_letter(3)].width = 18  # Last Name

    # Date columns - set uniform width for date columns
    for i, _ in enumerate(date_headers, 4):
        attendance_sheet.column_dimensions[get_column_letter(i)].width = 8

    # Add headers with formatting
    for col_num, header_text in enumerate(attendance_headers, 1):
        cell = attendance_sheet.cell(row=attendance_header_row, column=col_num, value=header_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = horizontal_border

    # Add employee attendance data starting from the row after headers
    row_num = attendance_header_row + 1

    # Add data for each employee
    for i, (emp_id, data) in enumerate(employee_attendance.items(), 1):
        # Add employee base info
        id_cell = attendance_sheet.cell(row=row_num, column=1, value=i)
        id_cell.border = horizontal_border
        id_cell.alignment = Alignment(horizontal='center', vertical='center')

        first_name_cell = attendance_sheet.cell(row=row_num, column=2, value=data["first_name"])
        first_name_cell.border = horizontal_border
        first_name_cell.alignment = Alignment(horizontal='left', vertical='center')

        last_name_cell = attendance_sheet.cell(row=row_num, column=3, value=data["last_name"])
        last_name_cell.border = horizontal_border
        last_name_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Add attendance status for each date
        for col_idx, current_date in enumerate(date_list, 4):
            # Get morning and afternoon status for this date
            am_pm_status = get_daily_attendance_status(data["punches"], current_date)

            # Create a cell with both statuses (morning/afternoon)
            cell_value = f"{am_pm_status[0]}\n{am_pm_status[1]}"
            cell = attendance_sheet.cell(row=row_num, column=col_idx, value=cell_value)

            # Format the cell
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = horizontal_border

        # Make the row taller to accommodate two lines of text
        attendance_sheet.row_dimensions[row_num].height = 30
        row_num += 1

    # Make the attendance sheet active when opening the file
    workbook.active = 0

    # Save the workbook
    full_path = os.path.join(report_directory, filename)
    workbook.save(full_path)

    # Try to open the Excel file automatically
    try:
        os.startfile(full_path) if os.name == 'nt' else subprocess.call(['open', full_path])
    except Exception as e:
        print(f"Error opening Excel file: {e}")